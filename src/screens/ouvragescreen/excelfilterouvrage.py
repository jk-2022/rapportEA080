from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter



class OuvrageExcelExporter:
    def __init__(self, datas: list, output_path: str, projet_name: str):
        self.datas = datas
        self.output_path = f"{output_path}.xlsx"
        self.wb = Workbook()
        self.ws = self.wb.active
        title=f"Liste Ouvrages du projet {projet_name}"
        self.ws.title =str(title)

        # Définition des colonnes
        self.ouvrage_fields = [
            "type_ouvrage", "numero_irh", "lieu", "localite", 
            "canton", "commune", "coordonnee_x", "coordonnee_y",
            "entreprise", "annee", "etat",
        ]

        self.foration_fields = [
            "date_foration", "prof_alteration",
            "prof_socle", "prof_total",
            "prof_tube_crepine", "prof_tube_plein",
            "debit_soufflage"
        ]

        self.pompage_fields = [
            "date_pompage", "type_pompe",
            "cote_pompe", "temps_pompage",
            "debit_pompage", "niv_dynamique",
            "niv_statique"
        ]

        # Couleurs
        self.ouvrage_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        self.foration_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        self.pompage_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    # ------------------------------------------------------

    def create_titles(self):
        ws = self.ws

        total_ouvrage = len(self.ouvrage_fields)
        total_foration = len(self.foration_fields)
        total_pompage = len(self.pompage_fields)

        col1 = 1
        col2 = col1 + total_ouvrage
        col3 = col2 + total_foration

        # Fusion OUVRAGES
        ws.merge_cells(start_row=1, start_column=col1,
                       end_row=1, end_column=col2 - 1)
        cell=ws.cell(row=1, column=col1, value="OUVRAGES")
        cell.fill = self.ouvrage_fill

        # Fusion FORATION
        ws.merge_cells(start_row=1, start_column=col2,
                       end_row=1, end_column=col3 - 1)
        cell=ws.cell(row=1, column=col2, value="FORATION")
        cell.fill = self.foration_fill
        
        # Fusion POMPAGE
        ws.merge_cells(start_row=1, start_column=col3,
                       end_row=1, end_column=col3 + total_pompage - 1)
        cell=ws.cell(row=1, column=col3, value="POMPAGE")
        cell.fill = self.pompage_fill
        
        # Style
        for col in range(1, col3 + total_pompage):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    # ------------------------------------------------------

    def create_headers(self):
        ws = self.ws
        headers = (
            self.ouvrage_fields +
            self.foration_fields +
            self.pompage_fields
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header.replace("_", " ").title()
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # ------------------------------------------------------

    def insert_data(self):
        ws = self.ws
        row_index = 3

        for item in self.datas:

            ouvrage = item.get("ouvrages", [{}])[0]
            foration = item.get("foration", {})
            pompage = item.get("pompage", {})

            row_data = []

            # Ouvrage
            for field in self.ouvrage_fields:
                row_data.append(ouvrage.get(field))

            # Foration
            for field in self.foration_fields:
                row_data.append(foration.get(field))

            # Pompage
            for field in self.pompage_fields:
                row_data.append(pompage.get(field))

            # Écriture ligne
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_index, column=col, value=value)

            row_index += 1

    # ------------------------------------------------------

    def auto_size_columns(self):
        ws = self.ws
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    # ------------------------------------------------------

    def add_borders(self):
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in self.ws.iter_rows():
            for cell in row:
                cell.border = border

    # ------------------------------------------------------

    def export(self):
        self.create_titles()
        self.create_headers()
        self.insert_data()
        self.auto_size_columns()
        self.add_borders()
        self.wb.save(self.output_path)
        return True
