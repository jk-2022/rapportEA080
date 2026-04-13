"""
services/export_service.py  — Exports xlsx / PDF pour EaRapport (structure réelle)
"""
import os, datetime
from typing import Optional
from myaction.db_actions import *
from mystorage import get_value

def get_archive_path():
    ARCHIVES_PATH=get_value("archive_path")
    return ARCHIVES_PATH

ARCHIVE_DIR=get_archive_path()

def _ts(): return datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
def _now(): return datetime.datetime.now().strftime("%d/%m/%Y à %H:%M")


# ─── LISTE OUVRAGES XLSX ──────────────────────────────────────────────────────

def export_ouvrages_xlsx(ouvrages, filename=None, projets=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Ouvrages"
    HF = PatternFill("solid", fgColor="1565C0")
    HFONT = Font(color="FFFFFF", bold=True, size=10)
    ODD = PatternFill("solid", fgColor="E3F2FD")
    BD = Border(left=Side(style="thin",color="BBDEFB"), right=Side(style="thin",color="BBDEFB"),
                top=Side(style="thin",color="BBDEFB"), bottom=Side(style="thin",color="BBDEFB"))
    AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
    headers = ["Préfecture","Commune","Canton","Localité","Lieu","X","Y",
               "Type","N° IRH","Année","Entreprise","Énergie","Réservoir","Volume",
               "État","Cause panne","Projet","Observation"]
    ws.append(headers)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(1, ci); c.fill=HF; c.font=HFONT; c.alignment=AC; c.border=BD
    def _pnom(o):
        if projets and o.projet_id:
            p = projets.get(o.projet_id); return p.name if p else ""
        return ""
    for i, o in enumerate(ouvrages, 2):
        row = [o.prefecture, o.commune, o.canton, o.localite, o.lieu,
               o.coordonnee_x, o.coordonnee_y,
               o.type_ouvrage, o.numero_irh, o.annee, o.entreprise,
               o.type_energie, o.type_reservoir, o.volume_reservoir,
               o.etat, o.cause_panne, _pnom(o), o.observation]
        ws.append(row)
        if i % 2 == 0:
            for ci in range(1, len(headers)+1): ws.cell(i,ci).fill = ODD
        for ci in range(1, len(headers)+1):
            ws.cell(i,ci).border = BD; ws.cell(i,ci).alignment = AC
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(mx+3, 35)
    path = os.path.join(ARCHIVE_DIR, filename or f"ouvrages_{_ts()}.xlsx")
    wb.save(path); return path


# ─── LISTE OUVRAGES PDF ───────────────────────────────────────────────────────

def export_ouvrages_pdf(ouvrages, filename=None, titre="Liste des Ouvrages", projets=None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER
    styles = getSampleStyleSheet()
    S_TITLE = ParagraphStyle("T", fontSize=14, textColor=colors.white,
                             backColor=colors.HexColor("#1565C0"),
                             fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=8)
    S_FOOT  = ParagraphStyle("F", fontSize=8, textColor=colors.HexColor("#607D8B"),
                             alignment=TA_CENTER)
    path = os.path.join(ARCHIVE_DIR, filename or f"ouvrages_{_ts()}.pdf")
    doc  = SimpleDocTemplate(path, pagesize=landscape(A4),
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=2*cm, bottomMargin=1.5*cm)
    def _pnom(o):
        if projets and o.projet_id:
            p = projets.get(o.projet_id); return p.name if p else ""
        return ""
    story = [Paragraph(f"EaRapport — {titre}", S_TITLE), Spacer(1, 0.3*cm)]
    headers = ["Préfecture","Commune","Localité","Type","État","Année","Entreprise","Projet"]
    data = [headers]
    for o in ouvrages:
        data.append([o.prefecture, o.commune, o.localite, o.type_ouvrage,
                     o.etat, o.annee, o.entreprise, _pnom(o)])
    cw = [3*cm,3*cm,4*cm,2*cm,2.5*cm,2*cm,5*cm,5*cm]
    tbl = Table(data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1565C0")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#E3F2FD")]),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#BBDEFB")),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("PADDING",(0,0),(-1,-1),4),
    ]))
    story += [tbl, Spacer(1,0.3*cm),
              Paragraph(f"Généré le {_now()} — {len(ouvrages)} ouvrage(s)", S_FOOT)]
    doc.build(story); return path


# ─── PV XLSX ──────────────────────────────────────────────────────────────────

def export_pv_xlsx(ouvrage, 
                   foration, 
                   pompage, 
                #    pannes, 
                #    suivis, 
                   projet_nom=""
                   ):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "PV"
    BLUE = PatternFill("solid", fgColor="1565C0")
    LIGHT= PatternFill("solid", fgColor="E3F2FD")
    WFONT= Font(color="FFFFFF", bold=True)
    BFONT= Font(bold=True)
    BD   = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))
    def _hdr(txt):
        ws.merge_cells(f"A{ws.max_row+1}:B{ws.max_row+1}")
        r = ws.max_row
        ws.cell(r,1).value=txt; ws.cell(r,1).fill=BLUE
        ws.cell(r,1).font=WFONT; ws.cell(r,1).border=BD
    def _row(lab, val, odd=False):
        ws.append([lab, val]); r=ws.max_row
        ws.cell(r,1).font=BFONT; ws.cell(r,1).border=BD; ws.cell(r,2).border=BD
        if odd: ws.cell(r,1).fill=LIGHT; ws.cell(r,2).fill=LIGHT
    o = ouvrage
    _hdr("PROCÈS-VERBAL D'OUVRAGE — " + (o.localite or o.type_ouvrage or ""))
    ws.cell(ws.max_row,1).font = Font(color="FFFFFF", bold=True, size=13)
    ws.append([])
    _hdr("IDENTIFICATION")
    for i,(l,v) in enumerate([
        ("Projet", projet_nom), ("Type d'ouvrage", o.type_ouvrage),
        ("État", o.etat), ("Année", o.annee),
        ("Entreprise", o.entreprise), ("N° IRH", o.numero_irh),
        ("Type énergie", o.type_energie), ("Type réservoir", o.type_reservoir),
        ("Volume réservoir", o.volume_reservoir), ("Cause panne", o.cause_panne),
    ]): _row(l, v, i%2==0)
    _hdr("LOCALISATION")
    for i,(l,v) in enumerate([
        ("Préfecture", o.prefecture), ("Commune", o.commune),
        ("Canton", o.canton), ("Localité", o.localite),
        ("Lieu-dit", o.lieu), ("Coordonnée X", o.coordonnee_x),
        ("Coordonnée Y", o.coordonnee_y),
    ]): _row(l, v, i%2==0)
    if foration:
        _hdr("DONNÉES DE FORATION")
        f = foration
        for i,(l,v) in enumerate([
            ("Date foration",f.date_foration),("Prof. altération",f.prof_alteration),
            ("Prof. socle",f.prof_socle),("Prof. totale",f.prof_total),
            ("Prof. tube crépine",f.prof_tube_crepine),("Prof. tube plein",f.prof_tube_plein),
            ("Débit soufflage",f.debit_soufflage),
        ]): _row(l, v, i%2==0)
    if pompage:
        _hdr("DONNÉES DE POMPAGE")
        p = pompage
        for i,(l,v) in enumerate([
            ("Date pompage",p.date_pompage),("Type pompe",p.type_pompe),
            ("Côte pompe",p.cote_pompe),("Temps pompage",p.temps_pompage),
            ("Débit pompage",p.debit_pompage),("Niveau dynamique",p.niv_dynamique),
            ("Niveau statique",p.niv_statique),
        ]): _row(l, v, i%2==0)
    # if pannes:
    #     ws.append([]); _hdr("PANNES")
    #     ws.append(["Date signalement","Description","Solution","Observation"])
    #     for p in pannes:
    #         ws.append([p.date_signaler, p.description, p.solution, p.observation])
    # if suivis:
    #     ws.append([]); _hdr("SUIVIS / RÉCEPTIONS")
    #     ws.append(["Date réception","Type","Participants","Recommandation","Observation"])
    #     for s in suivis:
    #         ws.append([s.date_reception, s.type_reception,
    #                    s.participants, s.recommandation, s.observation])
    ws.append([]); ws.append([f"Généré le {_now()}"])
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 45
    path = os.path.join(ARCHIVE_DIR, f"PV_{o.localite or o.type_ouvrage}_{_ts()}.xlsx")
    wb.save(path); return path


# ─── PV PDF ───────────────────────────────────────────────────────────────────

def _pdf_common_styles():
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("EaTitle", parent=styles["Title"],
        fontSize=15, textColor=colors.white, backColor=colors.HexColor("#1565C0"),
        spaceAfter=10, alignment=TA_CENTER))
    styles.add(ParagraphStyle("EaSection", parent=styles["Heading2"],
        fontSize=11, textColor=colors.white, backColor=colors.HexColor("#1E88E5"),
        spaceBefore=10, spaceAfter=5))
    styles.add(ParagraphStyle("EaBody", parent=styles["Normal"],
        fontSize=9, spaceAfter=3, leading=13))
    styles.add(ParagraphStyle("EaFoot", parent=styles["Normal"],
        fontSize=8, textColor=colors.HexColor("#607D8B"), alignment=TA_CENTER))
    return styles

def _two_col_table(pairs, col_w=None):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph
    styles = getSampleStyleSheet()
    SL = styles["Normal"]; SL.fontSize=8; SL.textColor=colors.HexColor("#546E7A")
    SV = styles["Normal"]; SV.fontSize=9
    col_w = col_w or [3.5*cm, 5*cm, 3.5*cm, 5*cm]
    rows = []
    for i in range(0, len(pairs), 2):
        l1, v1 = pairs[i]; l2, v2 = pairs[i+1] if i+1 < len(pairs) else ("","")
        rows.append([
            Paragraph(str(l1), SL), Paragraph(str(v1 or "—"), SV),
            Paragraph(str(l2), SL), Paragraph(str(v2 or "—"), SV),
        ])
    style = [
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#BBDEFB")),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
    ]
    for i in range(len(rows)):
        bg = colors.white if i%2==0 else colors.HexColor("#E3F2FD")
        style.append(("BACKGROUND",(0,i),(-1,i),bg))
    return Table(rows, colWidths=col_w, style=TableStyle(style))

def export_pv_pdf(ouvrage, foration, pompage, projet_nom=""):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.units import cm
    styles = _pdf_common_styles()
    o = ouvrage
    path = os.path.join(ARCHIVE_DIR, f"PV_{o.localite or o.type_ouvrage}_{_ts()}.pdf")
    doc  = SimpleDocTemplate(path, pagesize=A4,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []
    story.append(Paragraph(f"PROCÈS-VERBAL — {o.localite or ''} ({o.type_ouvrage})", styles["EaTitle"]))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph("Identification", styles["EaSection"]))
    story.append(_two_col_table([
        ("Projet", projet_nom), ("Type", o.type_ouvrage),
        ("État", o.etat), ("Année", o.annee),
        ("Entreprise", o.entreprise), ("N° IRH", o.numero_irh),
        ("Type énergie", o.type_energie), ("Type réservoir", o.type_reservoir),
        ("Volume réservoir", o.volume_reservoir), ("Cause panne", o.cause_panne),
    ]))
    story.append(Paragraph("Localisation", styles["EaSection"]))
    story.append(_two_col_table([
        ("Préfecture", o.prefecture), ("Commune", o.commune),
        ("Canton", o.canton), ("Localité", o.localite),
        ("Lieu-dit", o.lieu), ("Coordonnée X", o.coordonnee_x),
        ("Coordonnée Y", o.coordonnee_y), ("", ""),
    ]))
    if foration:
        f = foration
        story.append(Paragraph("Données de Foration", styles["EaSection"]))
        story.append(_two_col_table([
            ("Date foration", f.date_foration), ("Débit soufflage", f.debit_soufflage),
            ("Prof. altération", f.prof_alteration), ("Prof. socle", f.prof_socle),
            ("Prof. totale", f.prof_total), ("Prof. tube crépine", f.prof_tube_crepine),
            ("Prof. tube plein", f.prof_tube_plein), ("Observation", f.observation),
        ]))
    if pompage:
        p = pompage
        story.append(Paragraph("Données de Pompage", styles["EaSection"]))
        story.append(_two_col_table([
            ("Date pompage", p.date_pompage), ("Type pompe", p.type_pompe),
            ("Côte pompe", p.cote_pompe), ("Temps pompage", p.temps_pompage),
            ("Débit pompage", p.debit_pompage), ("Niv. dynamique", p.niv_dynamique),
            ("Niv. statique", p.niv_statique), ("Observation", p.observation),
        ]))
    # if pannes:
    #     story.append(Paragraph("Pannes", styles["EaSection"]))
    #     ph = [["Date","Description","Solution","Observation"]]
    #     for p in pannes: ph.append([p.date_signaler or "—", p.description or "—",
    #                                  p.solution or "—", p.observation or "—"])
    #     pt = Table(ph, colWidths=[2.5*cm,5.5*cm,5.5*cm,3.5*cm], repeatRows=1)
    #     pt.setStyle(TableStyle([
    #         ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1565C0")),
    #         ("TEXTCOLOR",(0,0),(-1,0),colors.white),
    #         ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
    #         ("FONTSIZE",(0,0),(-1,-1),8),
    #         ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#FFF8E1")]),
    #         ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#BBDEFB")),
    #         ("PADDING",(0,0),(-1,-1),4),("VALIGN",(0,0),(-1,-1),"TOP"),
    #     ]))
    #     story.append(pt)
    # if suivis:
    #     story.append(Paragraph("Suivis / Réceptions", styles["EaSection"]))
    #     sh = [["Date","Type","Participants","Recommandation","Obs."]]
    #     for s in suivis: sh.append([s.date_reception or "—", s.type_reception or "—",
    #                                  s.participants or "—", s.recommandation or "—",
    #                                  s.observation or "—"])
    #     st = Table(sh, colWidths=[2.5*cm,2.5*cm,4*cm,5*cm,3*cm], repeatRows=1)
    #     st.setStyle(TableStyle([
    #         ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#00ACC1")),
    #         ("TEXTCOLOR",(0,0),(-1,0),colors.white),
    #         ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
    #         ("FONTSIZE",(0,0),(-1,-1),8),
    #         ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#E0F7FA")]),
    #         ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#BBDEFB")),
    #         ("PADDING",(0,0),(-1,-1),4),("VALIGN",(0,0),(-1,-1),"TOP"),
    #     ]))
    #     story.append(st)
    story.append(Spacer(1,0.5*cm))
    story.append(HRFlowable(width="100%", thickness=0.8, color=colors.HexColor("#BBDEFB")))
    story.append(Paragraph(f"EaRapport — Généré le {_now()}", styles["EaFoot"]))
    doc.build(story); return path


# ─── RAPPORT DÉTAILLÉ PDF ─────────────────────────────────────────────────────

def export_rapport_detaille_pdf(ouvrage, foration=None, pompage=None,
                                pannes=None, suivis=None, projet_nom=""):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable, PageBreak)
    pannes = pannes or []; suivis = suivis or []
    o = ouvrage
    BLUE    = colors.HexColor("#1565C0"); BDARK = colors.HexColor("#0D47A1")
    BLIGHT  = colors.HexColor("#E3F2FD"); BPALE = colors.HexColor("#BBDEFB")
    CYAN    = colors.HexColor("#00ACC1")
    GREEN   = colors.HexColor("#2E7D32"); GBKG  = colors.HexColor("#E8F5E9")
    ORANGE  = colors.HexColor("#E65100"); OBKG  = colors.HexColor("#FFF3E0")
    RED     = colors.HexColor("#B71C1C")
    GREY    = colors.HexColor("#546E7A")
    WHITE   = colors.white
    ETAT_C  = {"Bon":(GREEN,GBKG),"En panne":(ORANGE,OBKG),"Abandonnée":(RED,colors.HexColor("#FFEBEE"))}
    etat_fg, etat_bg = ETAT_C.get(o.etat, (GREY, BLIGHT))

    SS = getSampleStyleSheet()
    def S(name, **kw): return ParagraphStyle(name, **kw)
    S_APPTITLE = S("AT", fontSize=9, textColor=colors.HexColor("#BBDEFB"),
                   fontName="Helvetica-Bold", alignment=TA_LEFT)
    S_DOCTYPE   = S("DT", fontSize=9, textColor=WHITE,
                   fontName="Helvetica-Bold", alignment=TA_RIGHT)
    S_CTITLE    = S("CT", fontSize=22, textColor=WHITE, fontName="Helvetica-Bold",
                   alignment=TA_CENTER, spaceAfter=6)
    S_CSUB      = S("CS", fontSize=12, textColor=colors.HexColor("#BBDEFB"),
                   fontName="Helvetica", alignment=TA_CENTER, spaceAfter=4)
    S_BADGE     = S("CB", fontSize=11, textColor=WHITE, fontName="Helvetica-Bold",
                   alignment=TA_CENTER)
    S_SECTION   = S("SEC", fontSize=11, textColor=WHITE, fontName="Helvetica-Bold",
                   backColor=BLUE, spaceBefore=12, spaceAfter=5)
    S_LABEL     = S("LAB", fontSize=8, textColor=GREY, fontName="Helvetica-Bold", spaceAfter=1)
    S_VALUE     = S("VAL", fontSize=9, textColor=colors.HexColor("#1A237E"),
                   fontName="Helvetica", spaceAfter=4)
    S_BODY      = S("BOD", fontSize=9, textColor=colors.HexColor("#1A237E"),
                   fontName="Helvetica", spaceAfter=3, leading=13)
    S_FOOT      = S("FT",  fontSize=8, textColor=GREY, alignment=TA_CENTER)
    S_NOTE      = S("NT",  fontSize=9, textColor=GREY, fontName="Helvetica-Oblique")

    def _hr(): return HRFlowable(width="100%",thickness=0.8,color=BPALE,spaceAfter=4,spaceBefore=4)
    def _section(txt):
        return [Spacer(1,0.25*cm),
                Table([[Paragraph(f"  {txt}", S_SECTION)]], colWidths=[17*cm],
                      style=TableStyle([("BACKGROUND",(0,0),(-1,-1),BLUE),
                                        ("ROWHEIGHT",(0,0),(-1,-1),22),
                                        ("LEFTPADDING",(0,0),(-1,-1),8),
                                        ("TOPPADDING",(0,0),(-1,-1),3),
                                        ("BOTTOMPADDING",(0,0),(-1,-1),3)])),
                Spacer(1,0.1*cm)]
    def _two(pairs):
        rows=[]
        for i in range(0,len(pairs),2):
            l1,v1=pairs[i]; l2,v2=pairs[i+1] if i+1<len(pairs) else ("","")
            rows.append([Paragraph(str(l1),S_LABEL),Paragraph(str(v1 or "—"),S_VALUE),
                         Paragraph(str(l2),S_LABEL),Paragraph(str(v2 or "—"),S_VALUE)])
        st=[("GRID",(0,0),(-1,-1),0.3,BPALE),("VALIGN",(0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
            ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)]
        for i in range(len(rows)):
            st.append(("BACKGROUND",(0,i),(-1,i),WHITE if i%2==0 else BLIGHT))
        return Table(rows,colWidths=[3.5*cm,5*cm,3.5*cm,5*cm],style=TableStyle(st))

    story=[]
    # En-tête bandeau
    story.append(Table([[Paragraph("EaRapport",S_APPTITLE),
                         Paragraph("RAPPORT DÉTAILLÉ D'OUVRAGE",S_DOCTYPE)]],
        colWidths=[8.5*cm,8.5*cm],
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),BDARK),
                           ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
                           ("LEFTPADDING",(0,0),(0,-1),10),("RIGHTPADDING",(-1,0),(-1,-1),10)])))
    story.append(Spacer(1,1*cm))
    story.append(Paragraph(f"{o.localite or '—'} — {o.commune}", S_CTITLE))
    story.append(Paragraph(f"Type : {o.type_ouvrage}", S_CSUB))
    story.append(Spacer(1,0.3*cm))
    story.append(Table([[Paragraph(f"  État : {o.etat}  ",S_BADGE)]],
        colWidths=[5*cm], hAlign="CENTER",
        style=TableStyle([("BACKGROUND",(0,0),(-1,-1),etat_fg),
                           ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)])))
    story.append(Spacer(1,0.7*cm)); story.append(_hr()); story.append(Spacer(1,0.2*cm))
    story.append(_two([("Projet",projet_nom),("Entreprise",o.entreprise),
                        ("Préfecture",o.prefecture),("Commune",o.commune),
                        ("Année",o.annee),("N° IRH",o.numero_irh)]))
    story.append(Spacer(1,0.5*cm)); story.append(_hr())
    story.append(Paragraph(f"Rapport généré le {_now()}  •  EaRapport v1.0.0", S_FOOT))
    story.append(PageBreak())

    story.extend(_section("1. Identification"))
    story.append(_two([("Type d'ouvrage",o.type_ouvrage),("État",o.etat),
                        ("N° IRH",o.numero_irh),("Année",o.annee),
                        ("Entreprise",o.entreprise),("Projet",projet_nom),
                        ("Type énergie",o.type_energie),("Type réservoir",o.type_reservoir),
                        ("Volume réservoir",o.volume_reservoir),("Cause panne",o.cause_panne)]))

    story.extend(_section("2. Localisation"))
    story.append(_two([("Préfecture",o.prefecture),("Commune",o.commune),
                        ("Canton",o.canton),("Localité",o.localite),
                        ("Lieu-dit",o.lieu),("Coordonnée X",o.coordonnee_x),
                        ("Coordonnée Y",o.coordonnee_y),("","")]))

    if o.observation:
        story.extend(_section("3. Observations"))
        story.append(Table([[Paragraph(o.observation or "—",S_BODY)]],colWidths=[17*cm],
            style=TableStyle([("BACKGROUND",(0,0),(-1,-1),BLIGHT),
                               ("GRID",(0,0),(-1,-1),0.3,BPALE),
                               ("PADDING",(0,0),(-1,-1),8)])))

    if foration:
        f=foration
        story.extend(_section("4. Données de Foration"))
        story.append(_two([("Date foration",f.date_foration),("Débit soufflage",f.debit_soufflage),
                            ("Prof. altération",f.prof_alteration),("Prof. socle",f.prof_socle),
                            ("Prof. totale",f.prof_total),("Prof. tube crépine",f.prof_tube_crepine),
                            ("Prof. tube plein",f.prof_tube_plein),("Observation",f.observation)]))

    if pompage:
        p=pompage
        story.extend(_section("5. Données de Pompage"))
        story.append(_two([("Date pompage",p.date_pompage),("Type pompe",p.type_pompe),
                            ("Côte pompe",p.cote_pompe),("Temps pompage",p.temps_pompage),
                            ("Débit pompage",p.debit_pompage),("Niv. dynamique",p.niv_dynamique),
                            ("Niv. statique",p.niv_statique),("Observation",p.observation)]))

    story.extend(_section(f"6. Pannes ({len(pannes)})"))
    if not pannes:
        story.append(Paragraph("Aucune panne enregistrée.", S_NOTE))
    else:
        ph=[["#","Date signalement","Description","Solution","Observation"]]
        for i,p in enumerate(pannes,1):
            ph.append([str(i),p.date_signaler or "—",p.description or "—",
                       p.solution or "—",p.observation or "—"])
        pt=Table(ph,colWidths=[0.7*cm,2.5*cm,5*cm,5*cm,4*cm],repeatRows=1)
        pt.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),BLUE),("TEXTCOLOR",(0,0),(-1,0),WHITE),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,OBKG]),
            ("GRID",(0,0),(-1,-1),0.4,BPALE),("PADDING",(0,0),(-1,-1),3),
            ("VALIGN",(0,0),(-1,-1),"TOP"),]))
        story.append(pt)

    story.extend(_section(f"7. Suivis / Réceptions ({len(suivis)})"))
    if not suivis:
        story.append(Paragraph("Aucun suivi enregistré.", S_NOTE))
    else:
        sh=[["#","Date","Type","Participants","Recommandation","Observation"]]
        for i,s in enumerate(suivis,1):
            sh.append([str(i),s.date_reception or "—",s.type_reception or "—",
                       s.participants or "—",s.recommandation or "—",s.observation or "—"])
        st=Table(sh,colWidths=[0.7*cm,2.5*cm,2.5*cm,4*cm,4*cm,3*cm],repeatRows=1)
        st.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),CYAN),("TEXTCOLOR",(0,0),(-1,0),WHITE),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,BLIGHT]),
            ("GRID",(0,0),(-1,-1),0.4,BPALE),("PADDING",(0,0),(-1,-1),3),
            ("VALIGN",(0,0),(-1,-1),"TOP"),]))
        story.append(st)

    story.append(Spacer(1,1*cm)); story.append(_hr())
    story.append(Paragraph(f"EaRapport — Rapport Détaillé — {o.localite or o.type_ouvrage}  •  {_now()}", S_FOOT))

    path = os.path.join(ARCHIVE_DIR, f"Rapport_{o.localite or o.type_ouvrage}_{_ts()}.pdf")
    doc  = SimpleDocTemplate(path, pagesize=A4,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=1.5*cm, bottomMargin=1.5*cm,
                             title=f"Rapport — {o.localite}", author="EaRapport")
    doc.build(story); return path


# ─── BACKUP JSON ─────────────────────────────────────────────────────────────

def export_backup_json(appstate):
    import json
    data = {
        "projets":     [vars(p) for p in appstate.load_projets()],
        "entreprises": [vars(e) for e in appstate.load_entreprises()],
        "villages":    [vars(v) for v in appstate.load_villages()],
        "ouvrages":    [vars(o) for o in appstate.load_all_ouvrages_flat()],
        "forations":    [vars(o) for o in appstate.load_forations()],
        "pompages":    [vars(o) for o in appstate.load_pompages()],
        "suivis":    [vars(o) for o in load_all_suivis()],
        "pannes":    [vars(o) for o in load_all_pannes()],
    }
    path = os.path.join(ARCHIVE_DIR, f"backup_{_ts()}.json")
    with open(path,"w",encoding="utf-8") as f: json.dump(data,f,ensure_ascii=False,indent=2)
    return path


# ─── BACKUP XLSX ─────────────────────────────────────────────────────────────

def export_backup_xlsx(appstate):
    from openpyxl import Workbook
    wb = Workbook()
    def _ws(name, items, fields):
        ws = wb.create_sheet(name); ws.append(fields)
        for item in items: ws.append([getattr(item,f,"") for f in fields])
    _ws("Projets",     appstate.load_projets(),
        ["id","name","title","secteurs"])
    _ws("Entreprises", appstate.load_entreprises(),
        ["id","name","contact"])
    _ws("Villages",    appstate.load_villages(),
        ["id","prefecture","commune","canton","localite","coordonnee_x","coordonnee_y",
         "ressource","status","observation"])
    _ws("Ouvrages",    appstate.load_all_ouvrages_flat(),
        ["id","projet_id","prefecture","commune","canton","localite","lieu",
         "coordonnee_x","coordonnee_y","entreprise","type_ouvrage","numero_irh","annee",
         "type_energie","type_reservoir","volume_reservoir","etat","cause_panne","observation"])
    _ws("Forations",    appstate.load_forations(),
        ["id","ouvrage_id", "date_foration", "prof_alteration", "prof_socle", "prof_total", "prof_tube_crepine", "prof_tube_plein", "debit_soufflage", "observation"])
    _ws("Pompages",    appstate.load_pompages(),
        ["id","ouvrage_id", "date_pompage", "type_pompe", "cote_pompe", "temps_pompage", "debit_pompage", "niv_dynamique", "niv_statique", "observation"])
    _ws("Pannes",    load_all_pannes(),
        ["id","ouvrage_id", "date_signaler", "description", "solution", "observation"])
    _ws("Suivis",    load_all_suivis(),
        ["id","ouvrage_id", "date_reception", "type_reception", "participants", "recommandation", "observation"])
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    path = os.path.join(ARCHIVE_DIR, f"backup_{_ts()}.xlsx")
    wb.save(path); return path


# ─── IMPORT JSON ─────────────────────────────────────────────────────────────

def import_backup_json(path, appstate):
    import json
    # from database.models import (save_projet, save_entreprise, save_village, save_ouvrage,
    #                               Projet, Entreprise, Village, Ouvrage)
    with open(path,"r",encoding="utf-8") as f: data=json.load(f)
    def _clean(cls, item):
        item.pop("id",None); item.pop("created_at",None)
        fields = set(cls.__dataclass_fields__)
        return cls(**{k:v for k,v in item.items() if k in fields})
    for item in data.get("projets",[]): create_projet(_clean(Projet,item))
    for item in data.get("entreprises",[]): create_entreprise(_clean(Entreprise,item))
    for item in data.get("villages",[]): create_village(_clean(Village,item))
    for item in data.get("ouvrages",[]): create_ouvrage(_clean(Ouvrage,item))
    for item in data.get("forations",[]): create_foration(_clean(Ouvrage,item))
    for item in data.get("pompages",[]): create_pompage(_clean(Ouvrage,item))
    for item in data.get("pannes",[]): create_panne(_clean(Ouvrage,item))
    for item in data.get("suivi",[]): create_suivi(_clean(Ouvrage,item))
    # appstate.initialize()


# ─── IMPORT XLSX ─────────────────────────────────────────────────────────────

def import_backup_xlsx(path, appstate):
    from openpyxl import load_workbook

    wb   = load_workbook(path, read_only=True, data_only=True)
    conn = connected_db()
    errors = []

    def _rows(sheet_name):
        """Lit une feuille et retourne une liste de dicts colonne→valeur."""
        if sheet_name not in wb.sheetnames:
            return []
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [
            {headers[i]: (row[i] if i < len(row) else None)
             for i in range(len(headers))}
            for row in rows[1:]
            if any(v is not None for v in row)
        ]

    def _s(v): return str(v).strip() if v is not None else ""
    def _f(v):
        try: return float(v) if v is not None else 0.0
        except: return 0.0
    def _i(v):
        """Entier ou None."""
        try: return int(v) if v is not None else None
        except: return None

    # ── Projets ──────────────────────────────────────────────────────────────
    for r in _rows("Projets"):
        try:
            conn.execute("""
                INSERT OR REPLACE INTO projets (id, name, title, secteurs, created_at)
                VALUES (?,?,?,?,?)
            """, (_i(r.get("id")), _s(r.get("name")), _s(r.get("title")),
                  _s(r.get("secteurs")), _s(r.get("created_at"))))
        except Exception as e:
            errors.append(f"Projet id={r.get('id')}: {e}")

    # ── Entreprises ───────────────────────────────────────────────────────────
    for r in _rows("Entreprises"):
        try:
            conn.execute("""
                INSERT OR REPLACE INTO entreprises (id, name, contact, created_at)
                VALUES (?,?,?,?)
            """, (_i(r.get("id")), _s(r.get("name")),
                  _s(r.get("contact")), _s(r.get("created_at"))))
        except Exception as e:
            errors.append(f"Entreprise id={r.get('id')}: {e}")

    # ── Villages ──────────────────────────────────────────────────────────────
    for r in _rows("Villages"):
        try:
            conn.execute("""
                INSERT OR REPLACE INTO villages
                (id, prefecture, commune, canton, localite,
                 coordonnee_x, coordonnee_y, ressource, status, observation, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (_i(r.get("id")), _s(r.get("prefecture")), _s(r.get("commune")),
                  _s(r.get("canton")), _s(r.get("localite")),
                  _f(r.get("coordonnee_x")), _f(r.get("coordonnee_y")),
                  _s(r.get("ressource")), _s(r.get("status")),
                  _s(r.get("observation")), _s(r.get("created_at"))))
        except Exception as e:
            errors.append(f"Village id={r.get('id')}: {e}")

    # ── Ouvrages ──────────────────────────────────────────────────────────────
    # projet_id est réutilisé tel quel — c'est l'id original du fichier,
    # qui correspond à l'id réel en base après restauration des projets ci-dessus.
    for r in _rows("Ouvrages"):
        try:
            conn.execute("""
                INSERT OR REPLACE INTO ouvrages
                (id, projet_id, prefecture, commune, canton, localite, lieu,
                 coordonnee_x, coordonnee_y, entreprise, type_ouvrage, numero_irh,
                 annee, type_energie, type_reservoir, volume_reservoir,
                 etat, cause_panne, observation, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (_i(r.get("id")), _i(r.get("projet_id")),
                  _s(r.get("prefecture")), _s(r.get("commune")),
                  _s(r.get("canton")), _s(r.get("localite")),
                  _s(r.get("lieu")),
                  _f(r.get("coordonnee_x")), _f(r.get("coordonnee_y")),
                  _s(r.get("entreprise")), _s(r.get("type_ouvrage")) or "PMH",
                  _f(r.get("numero_irh")), _s(r.get("annee")),
                  _s(r.get("type_energie")), _s(r.get("type_reservoir")),
                  _s(r.get("volume_reservoir")),
                  _s(r.get("etat")) or "Bon",
                  _s(r.get("cause_panne")), _s(r.get("observation")),
                  _s(r.get("created_at"))))
        except Exception as e:
            errors.append(f"Ouvrage id={r.get('id')}: {e}")
            
    
    # ── Forations ──────────────────────────────────────────────────────────────
    # ouvrage_id est réutilisé tel quel — c'est l'id original du fichier,
    # qui correspond à l'id réel en base après restauration des ouvrages ci-dessus.
    for r in _rows("Forations"):
        try:
            conn.execute("""
                INSERT OR REPLACE INTO foration
                (id, ouvrage_id, date_foration, prof_alteration, prof_socle, prof_total, prof_tube_crepine, prof_tube_plein, debit_soufflage, observation, "created_at") VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """, (_i(r.get("id")), _i(r.get("ouvrage_id")),
                  _s(r.get("date_foration")), _s(r.get("prof_alteration")),
                  _s(r.get("prof_socle")), _s(r.get("prof_total")),
                  _s(r.get("prof_tube_crepine")),
                  _f(r.get("prof_tube_plein")), _f(r.get("debit_soufflage")), _s(r.get("observation")),
                  _s(r.get("created_at"))))
        except Exception as e:
            errors.append(f"Foration id={r.get('id')}: {e}")
            
    # ── Pompages ──────────────────────────────────────────────────────────────
    # ouvrage_id est réutilisé tel quel — c'est l'id original du fichier,
    # qui correspond à l'id réel en base après restauration des ouvrages ci-dessus.
    for r in _rows("Pompages"):
        try:
            conn.execute("""
                INSERT OR REPLACE INTO pompage
                (id, ouvrage_id, date_pompage, type_pompe, cote_pompe, temps_pompage, debit_pompage, niv_dynamique, niv_statique, observation, "created_at") VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """, (_i(r.get("id")), _i(r.get("ouvrage_id")),
                  _s(r.get("date_pompage")), _s(r.get("type_pompe")),
                  _s(r.get("cote_pompe")), _s(r.get("temps_pompage")),
                  _s(r.get("debit_pompage")),
                  _f(r.get("niv_dynamique")), _f(r.get("niv_statique")), _s(r.get("observation")),
                  _s(r.get("created_at"))))
        except Exception as e:
            errors.append(f"Pompage id={r.get('id')}: {e}")

    # ── Pannes ──────────────────────────────────────────────────────────────
    for r in _rows("Pannes"):
        try:
            conn.execute("""
                INSERT OR REPLACE INTO panne
                (id, ouvrage_id, date_signaler, description, solution, observation, "created_at") VALUES(?,?,?,?,?,?,?)
            """, (_i(r.get("id")), _s(r.get("ouvrage_id")), _s(r.get("date_signaler")),
                    _s(r.get("description")), _s(r.get("solution")),
                    _s(r.get("observation")), _s(r.get("created_at"))))
        except Exception as e:
            errors.append(f"Panne id={r.get('id')}: {e}")
            
    
    # ── Suivis ──────────────────────────────────────────────────────────────
    for r in _rows("Suivis"):
        try:
            conn.execute("""
                INSERT OR REPLACE INTO suivi
                (id, ouvrage_id, date_reception, type_reception, participants, recommandation, observation,"created_at") VALUES(?,?,?,?,?,?,?,?)
            """, (_i(r.get("id")), _s(r.get("ouvrage_id")), _s(r.get("date_reception")),
                    _s(r.get("type_reception")), _s(r.get("participants")), _s(r.get("recommandation")),
                    _s(r.get("observation")), _s(r.get("created_at"))))
        except Exception as e:
            errors.append(f"Suivi id={r.get('id')}: {e}")
            
    conn.commit()
    conn.close()
    wb.close()
    # appstate.initialize()
    return errors

            
            
# ─── ARCHIVES ────────────────────────────────────────────────────────────────

def list_archives():
    files=[]
    if os.path.exists(ARCHIVE_DIR):
        for f in sorted(os.listdir(ARCHIVE_DIR), reverse=True):
            full=os.path.join(ARCHIVE_DIR,f)
            if os.path.isfile(full):
                files.append({"name":f,"path":full,
                               "size":os.path.getsize(full),
                               "mtime":datetime.datetime.fromtimestamp(
                                   os.path.getmtime(full)).strftime("%d/%m/%Y %H:%M"),
                               "ext":os.path.splitext(f)[1].lower()})
    return files
